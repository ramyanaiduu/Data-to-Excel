import openpyxl

def add_student_data():
    students = []

    while True:
        print("Enter student details:")
        name = input("Name: ")
        roll_no = input("Roll No: ")
        branch = input("Branch: ")
        phone_no = input("Phone No: ")

        student = {
            "Name": name,
            "Roll No": roll_no,
            "Branch": branch,
            "Phone No": phone_no,
        }
        students.append(student)

        more_data = input("Do you want to enter data for another student? (yes/no): ")
        if more_data.lower() != "yes":
            break

    return students

def create_excel(students, output_file):
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Write headers
    headers = list(students[0].keys())
    for col_idx, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=col_idx, value=header)

    # Write student data
    for row_idx, student in enumerate(students, start=2):
        for col_idx, value in enumerate(student.values(), start=1):
            sheet.cell(row=row_idx, column=col_idx, value=value)

    # Save the Excel file
    workbook.save(output_file)

if _name_ == "_main_":
    print("Enter student data. Type 'done' when you finish.")

    student_data = add_student_data()

    if student_data:
        print("\nStudent data collected successfully.")
        output_file_name = input("Enter the output Excel file name: ")

        create_excel(student_data, output_file_name)
        print(f"Data successfully converted and saved to {output_file_name}.")
    else:
        print("No student data to convert.")